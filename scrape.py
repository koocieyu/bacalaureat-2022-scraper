import unicodedata
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook(filename="medii.xlsx")
foaie = wb["alfabetic_tara"]

def procesare(candidat):
    bg_green = PatternFill(start_color="B9F6CA", fill_type="solid")
    print(f"Procesez candidatul {candidat[1]} (#{candidat[0]})...", end="")
    ic = int(candidat[0])+1
    
    echivalabile = {
        9: "J", 10: "K", 12: "M", 15: "S", 18: "AG", 21: "O", 22: "P", 24: "R", 25: "Z", 27: "AB", 28: "AD", 30: "AF"
    }
    
    for i in range (9, 31):
        if candidat[i].endswith("@@"):
            candidat[i] = candidat[i][:-2]
            foaie[f"{echivalabile[i]}{ic}"].fill = bg_green
    
    date_excel = {
        0: "A", 1: "B", 2: "C", 3: "D", 4: "E", 5: "F", 6: "G", 7: "H", 8: "I", 9: "J", 10: "K", 11: "L",
        12: "M", 13: "N", 14: "S", 16: "Y", 17: "AC", 18: "AG", 19: "AH", 20: "AI",
        21: "O", 22: "P", 23: "Q", 24: "R", 25: "Z", 26: "AA", 27: "AB", 28: "AD", 29: "AE", 30: "AF",
    }
    
    for column in date_excel.keys():
        foaie[f"{date_excel[column]}{ic}"] = candidat[column]

    if candidat[15]:
        split_moderna = candidat[15].split("-")
        if len(split_moderna) != 5:
            split_moderna = [candidat[15], candidat[15], candidat[15], candidat[15], candidat[15]]
        foaie[f'T{ic}'] = split_moderna[0]
        foaie[f'U{ic}'] = split_moderna[1]
        foaie[f'V{ic}'] = split_moderna[2]
        foaie[f'W{ic}'] = split_moderna[3]
        foaie[f'X{ic}'] = split_moderna[4]

    print("gata!")
        
        

with sync_playwright() as p:
    browser = p.webkit.launch(headless=True)
    page=browser.new_page()
    for i in range(1, 12647):
        print(f"Incarc pagina de rezultate {i}/12646...", end="")
        page.goto(f"http://static.bacalaureat.edu.ro/2022/rapoarte/rezultate/alfabetic/page_{i}.html")
        page.wait_for_load_state("load")
        print("gata!")
        html_tabel = page.inner_html("#mainTable>tbody")

        soup = BeautifulSoup(html_tabel, "html.parser")
        data = soup.find_all("td")
        values = []
        index_max = 31
        index = 0

        for tabledata in data:
            x = BeautifulSoup(str(tabledata), "html.parser")

            if "bgcolor" in list(x.td.attrs):
                values.append(f"{unicodedata.normalize('NFKC', x.text).strip()}@@")
            else:
                values.append(unicodedata.normalize("NFKC", x.text).strip())

            index += 1
            if index == index_max:
                procesare(values)
                values.clear()
                index = 0
        if i % 50 == 0:
            wb.save("medii.xlsx")
        
    wb.save("medii.xlsx")
        # souped_data = BeautifulSoup(d, "html.parser")
        # data_new.append(souped_data.text)
        
        # if i == 9: foaie[f'J{ic}'].fill = bg_green
        # elif i == 10: foaie[f'K{ic}'].fill = bg_green
        # elif i == 12: foaie[f'M{ic}'].fill = bg_green
        # elif i == 15: foaie[f'S{ic}'].fill = bg_green # lm
        # elif i == 18: foaie[f'AG{ic}'].fill = bg_green
        # elif i == 21: foaie[f'O{ic}'].fill = bg_green
        # elif i == 22: foaie[f'P{ic}'].fill = bg_green
        # elif i == 24: foaie[f'R{ic}'].fill = bg_green
        # elif i == 25: foaie[f'Z{ic}'].fill = bg_green
        # elif i == 27: foaie[f'AB{ic}'].fill = bg_green
        # elif i == 28: foaie[f'AD{ic}'].fill = bg_green
        # elif i == 30: foaie[f'AF{ic}'].fill = bg_green
        
        # foaie[f'A{ic}'] = candidat[0]
        # foaie[f'B{ic}'] = candidat[1]
        # # foaie[f'A{ic}'].fill = bg_green
        # foaie[f'C{ic}'] = candidat[2]
        # foaie[f'D{ic}'] = candidat[3]
        # foaie[f'E{ic}'] = candidat[4]
        # foaie[f'F{ic}'] = candidat[5]
        # foaie[f'G{ic}'] = candidat[6]
        # foaie[f'H{ic}'] = candidat[7]
        # foaie[f'I{ic}'] = candidat[8]
        # foaie[f'J{ic}'] = candidat[9]
        # foaie[f'K{ic}'] = candidat[10]
        # foaie[f'L{ic}'] = candidat[11]
        # foaie[f'M{ic}'] = candidat[12]
        # foaie[f'N{ic}'] = candidat[13]
        # foaie[f'S{ic}'] = candidat[14]
        # foaie[f'Y{ic}'] = candidat[16]
        # foaie[f'AC{ic}'] = candidat[17]
        # foaie[f'AG{ic}'] = candidat[18]
        # foaie[f'AH{ic}'] = candidat[19]
        # foaie[f'AI{ic}'] = candidat[20]
        # foaie[f'O{ic}'] = candidat[21]
        # foaie[f'P{ic}'] = candidat[22]
        # foaie[f'Q{ic}'] = candidat[23]
        # foaie[f'R{ic}'] = candidat[24]
        # foaie[f'Z{ic}'] = candidat[25]
        # foaie[f'AA{ic}'] = candidat[26]
        # foaie[f'AB{ic}'] = candidat[27]
        # foaie[f'AD{ic}'] = candidat[28]
        # foaie[f'AE{ic}'] = candidat[29]
        # foaie[f'AF{ic}'] = candidat[30]
        # '', '', '', '', '3.2', '3.05', '3.05', '4.1', '3.6', '3.6'